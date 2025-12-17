import asyncio
import aiohttp
import csv
from io import BytesIO, StringIO
from typing import List, Dict, Any
import tiktoken
from openai import OpenAI
from pinecone import Pinecone, ServerlessSpec
from openpyxl import load_workbook

from config import settings
from db import database

# Initialize clients
client = OpenAI(api_key=settings.openai_api_key)
pc = Pinecone(api_key=settings.pinecone_api_key)


# -------------------- Pinecone --------------------

def get_or_create_index():
    """Get or create Pinecone index"""
    if settings.pinecone_index_name not in pc.list_indexes().names():
        pc.create_index(
            name=settings.pinecone_index_name,
            dimension=1536,
            metric="cosine",
            spec=ServerlessSpec(
                cloud="aws",
                region="us-east-1"
            )
        )
    return pc.Index(settings.pinecone_index_name)


# -------------------- Fetch --------------------

async def fetch_file(url: str) -> bytes:
    """Fetch file content from URL"""
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            if response.status != 200:
                raise Exception(f"Failed to fetch file: HTTP {response.status}")
            return await response.read()


# -------------------- Parse (NO PANDAS) --------------------

def parse_csv(content: bytes) -> List[Dict[str, Any]]:
    text = content.decode("utf-8")
    reader = csv.DictReader(StringIO(text))
    return [dict(row) for row in reader]


def parse_excel(content: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() for h in rows[0]]
    data = []

    for row in rows[1:]:
        record = {}
        for key, val in zip(headers, row):
            record[key] = val
        data.append(record)

    return data


def parse_file(content: bytes, url: str) -> List[Dict[str, Any]]:
    """Parse CSV or Excel file"""
    if url.endswith(".csv"):
        return parse_csv(content)
    elif url.endswith((".xlsx", ".xls")):
        return parse_excel(content)
    else:
        raise ValueError("Unsupported file format. Only CSV and Excel files are supported.")


# -------------------- Text Processing --------------------

def create_property_text(row: Dict[str, Any]) -> str:
    """Convert a property row to structured text"""
    parts = []
    for col, val in row.items():
        if val is not None and val != "":
            parts.append(f"{col}: {val}")
    return "\n".join(parts)


def chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> List[str]:
    encoding = tiktoken.get_encoding("cl100k_base")
    tokens = encoding.encode(text)

    chunks = []
    start = 0

    while start < len(tokens):
        end = start + chunk_size
        chunk_tokens = tokens[start:end]
        chunks.append(encoding.decode(chunk_tokens))
        start = end - overlap

    return chunks


# -------------------- Embeddings --------------------

def generate_embeddings(texts: List[str]) -> List[List[float]]:
    response = client.embeddings.create(
        model=settings.embedding_model,
        input=texts
    )
    return [item.embedding for item in response.data]


# -------------------- Processing --------------------

async def process_single_source(source_id: str, url: str):
    try:
        database.update_source_status(source_id, "processing")

        content = await fetch_file(url)
        rows = parse_file(content, url)

        index = get_or_create_index()
        vectors_to_upsert = []

        for idx, row in enumerate(rows):
            property_text = create_property_text(row)
            chunks = chunk_text(
                property_text,
                settings.chunk_size,
                settings.chunk_overlap
            )

            embeddings = generate_embeddings(chunks)

            for chunk_idx, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
                vector_id = f"{source_id}_{idx}_{chunk_idx}"

                metadata = {
                    "source_id": source_id,
                    "property_index": idx,
                    "chunk_index": chunk_idx,
                    "text": chunk,
                    "url": url
                }

                # Add row fields to metadata
                for col, val in row.items():
                    if val is not None and val != "":
                        key = str(col).lower().replace(" ", "_")
                        metadata[key] = str(val)

                vectors_to_upsert.append(
                    (vector_id, embedding, metadata)
                )

        # Batch upsert
        batch_size = 100
        for i in range(0, len(vectors_to_upsert), batch_size):
            index.upsert(vectors=vectors_to_upsert[i:i + batch_size])

        database.update_source_status(source_id, "completed")

    except Exception as e:
        database.update_source_status(source_id, "failed", str(e))
        print(f"Error processing source {source_id}: {e}")


# -------------------- Batch Runner --------------------

async def ingest_all_sources():
    sources = database.get_all_sources()
    if not sources:
        return

    tasks = [
        process_single_source(source["id"], source["url"])
        for source in sources
    ]

    await asyncio.gather(*tasks, return_exceptions=True)
